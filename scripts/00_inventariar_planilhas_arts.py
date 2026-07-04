# -*- coding: utf-8 -*-
"""Inventaria planilhas anuais de ARTs mantidas localmente.

Uso recomendado no PowerShell, a partir da raiz do repositorio:

    python scripts/00_inventariar_planilhas_arts.py --entrada data/local --saida relatorios/inventario_planilhas

Este script nao publica dados linha a linha. Ele gera apenas inventario tecnico
sobre arquivos, abas, cabecalhos e contagens basicas.
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
import zipfile
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, List, Optional

EXTENSOES_SUPORTADAS = {".csv", ".xlsx", ".xlsm"}
VALOR_INSUFICIENTE = "Informacao insuficiente para verificar."


@dataclass
class InventarioArquivo:
    arquivo: str
    extensao: str
    tamanho_bytes: int
    status: str
    abas: str = ""
    total_linhas_estimado: str = ""
    total_colunas_estimado: str = ""
    cabecalhos_detectados: str = ""
    campo_art_detectado: str = ""
    campo_municipio_detectado: str = ""
    campo_valor_detectado: str = ""
    campo_tos_detectado: str = ""
    observacao: str = ""


def normalizar_token(texto: object) -> str:
    return str(texto or "").strip().lower().replace("ç", "c").replace("ã", "a").replace("á", "a").replace("à", "a").replace("â", "a").replace("é", "e").replace("ê", "e").replace("í", "i").replace("ó", "o").replace("ô", "o").replace("õ", "o").replace("ú", "u")


def detectar_campos(cabecalhos: Iterable[object]) -> dict:
    campos = {"art": "", "municipio": "", "valor": "", "tos": ""}
    for cab in cabecalhos:
        original = str(cab or "").strip()
        c = normalizar_token(original)
        if not campos["art"] and any(t in c for t in ["art", "numero art", "n art", "id_art"]):
            campos["art"] = original
        if not campos["municipio"] and any(t in c for t in ["municipio", "cidade", "localidade"]):
            campos["municipio"] = original
        if not campos["valor"] and any(t in c for t in ["valor", "contrato", "honorario", "taxa"]):
            campos["valor"] = original
        if not campos["tos"] and any(t in c for t in ["tos", "codigo tos", "cod tos", "atividade tecnica"]):
            campos["tos"] = original
    return {k: (v or VALOR_INSUFICIENTE) for k, v in campos.items()}


def inventariar_csv(path: Path) -> InventarioArquivo:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            amostra = f.read(4096)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(amostra, delimiters=";,\t,")
            except csv.Error:
                dialect = csv.excel
            reader = csv.reader(f, dialect)
            cab = next(reader, [])
            linhas = 1 + sum(1 for _ in reader) if cab else 0
        campos = detectar_campos(cab)
        return InventarioArquivo(
            arquivo=str(path),
            extensao=path.suffix.lower(),
            tamanho_bytes=path.stat().st_size,
            status="ok",
            abas="CSV",
            total_linhas_estimado=str(linhas),
            total_colunas_estimado=str(len(cab)),
            cabecalhos_detectados=" | ".join(str(x) for x in cab),
            campo_art_detectado=campos["art"],
            campo_municipio_detectado=campos["municipio"],
            campo_valor_detectado=campos["valor"],
            campo_tos_detectado=campos["tos"],
        )
    except Exception as exc:  # noqa: BLE001
        return InventarioArquivo(str(path), path.suffix.lower(), path.stat().st_size, "erro", observacao=str(exc))


def inventariar_xlsx(path: Path) -> InventarioArquivo:
    try:
        try:
            import openpyxl  # type: ignore
        except ImportError:
            return InventarioArquivo(
                arquivo=str(path),
                extensao=path.suffix.lower(),
                tamanho_bytes=path.stat().st_size,
                status="dependencia_ausente",
                observacao="Instale openpyxl para ler cabecalhos de XLSX: pip install openpyxl",
            )

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        abas = wb.sheetnames
        primeira = wb[abas[0]] if abas else None
        cab = []
        linhas = ""
        colunas = ""
        if primeira is not None:
            cab = [cell.value for cell in next(primeira.iter_rows(min_row=1, max_row=1), [])]
            linhas = str(primeira.max_row or "")
            colunas = str(primeira.max_column or "")
        campos = detectar_campos(cab)
        wb.close()
        return InventarioArquivo(
            arquivo=str(path),
            extensao=path.suffix.lower(),
            tamanho_bytes=path.stat().st_size,
            status="ok",
            abas=" | ".join(abas),
            total_linhas_estimado=linhas,
            total_colunas_estimado=colunas,
            cabecalhos_detectados=" | ".join(str(x or "") for x in cab),
            campo_art_detectado=campos["art"],
            campo_municipio_detectado=campos["municipio"],
            campo_valor_detectado=campos["valor"],
            campo_tos_detectado=campos["tos"],
        )
    except zipfile.BadZipFile as exc:
        return InventarioArquivo(str(path), path.suffix.lower(), path.stat().st_size, "erro", observacao=f"XLSX invalido: {exc}")
    except Exception as exc:  # noqa: BLE001
        return InventarioArquivo(str(path), path.suffix.lower(), path.stat().st_size, "erro", observacao=str(exc))


def encontrar_arquivos(entrada: Path) -> List[Path]:
    if not entrada.exists():
        return []
    return sorted(p for p in entrada.rglob("*") if p.is_file() and p.suffix.lower() in EXTENSOES_SUPORTADAS)


def salvar_csv(registros: List[InventarioArquivo], destino: Path) -> None:
    destino.parent.mkdir(parents=True, exist_ok=True)
    with destino.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(asdict(registros[0]).keys()) if registros else list(InventarioArquivo("", "", 0, "").__dataclass_fields__.keys()))
        writer.writeheader()
        for r in registros:
            writer.writerow(asdict(r))


def salvar_md(registros: List[InventarioArquivo], destino: Path, entrada: Path) -> None:
    destino.parent.mkdir(parents=True, exist_ok=True)
    total = len(registros)
    ok = sum(1 for r in registros if r.status == "ok")
    pend = total - ok
    linhas = [
        "# Inventario de planilhas de ARTs",
        "",
        f"Gerado em: {datetime.now().isoformat(timespec='seconds')}",
        f"Entrada analisada: `{entrada}`",
        "",
        f"Arquivos encontrados: {total}",
        f"Arquivos lidos com sucesso: {ok}",
        f"Arquivos com pendencia: {pend}",
        "",
        "## Arquivos",
        "",
        "| Arquivo | Status | Abas | Linhas estimadas | Campo ART | Campo municipio | Campo valor | Campo TOS | Observacao |",
        "|---|---:|---|---:|---|---|---|---|---|",
    ]
    for r in registros:
        linhas.append(
            f"| `{r.arquivo}` | {r.status} | {r.abas} | {r.total_linhas_estimado} | {r.campo_art_detectado} | {r.campo_municipio_detectado} | {r.campo_valor_detectado} | {r.campo_tos_detectado} | {r.observacao} |"
        )
    if not registros:
        linhas.append("| Informação insuficiente para verificar. | - | - | - | - | - | - | - | Nenhum arquivo CSV/XLSX encontrado. |")
    destino.write_text("\n".join(linhas) + "\n", encoding="utf-8")


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Inventaria planilhas locais de ARTs sem publicar dados linha a linha.")
    parser.add_argument("--entrada", default="data/local", help="Pasta local com planilhas de ARTs.")
    parser.add_argument("--saida", default="relatorios/inventario_planilhas", help="Prefixo dos arquivos de saida sem extensao.")
    args = parser.parse_args(argv)

    entrada = Path(args.entrada)
    saida = Path(args.saida)
    arquivos = encontrar_arquivos(entrada)
    registros: List[InventarioArquivo] = []
    for arq in arquivos:
        if arq.suffix.lower() == ".csv":
            registros.append(inventariar_csv(arq))
        else:
            registros.append(inventariar_xlsx(arq))

    salvar_csv(registros, saida.with_suffix(".csv"))
    salvar_md(registros, saida.with_suffix(".md"), entrada)
    print(f"Inventario gerado: {saida.with_suffix('.csv')} e {saida.with_suffix('.md')}")
    if not arquivos:
        print("Informacao insuficiente para verificar: nenhuma planilha local encontrada.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
