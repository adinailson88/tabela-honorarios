# -*- coding: utf-8 -*-
"""Gera PLANILHA_MODELO_HONORARIOS.xlsx a partir da calibracao APARADA (faixas piso/ref/teto + confiabilidade).
NAO inventa honorarios: honorario-referencia fica 'Informacao insuficiente para verificar' ate a pesquisa de precos.
Inclui aba de precificabilidade (o que NAO da para precificar por falta de valor/unidade)."""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE=os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
CAL=os.path.join(BASE,'dados','calibracao_atividade_unidade.csv')
PREC=os.path.join(BASE,'dados','precificabilidade_por_tipo.csv')

def grupo(ativ,unid):
    a=ativ.lower(); u=unid.lower()
    if 'pico' in u or 'quilowatt' in u or 'volt' in u: return 'Elétrica / Energia'
    if 'hectare' in u: return 'Agronomia / Rural'
    if 'metro cúbico' in u: return 'Hidráulica / Saneamento'
    if 'quilômetro' in u or u=='metro': return 'Obras lineares / Infraestrutura'
    if 'projeto' in a or 'orçamento' in a: return 'Projeto'
    if 'laudo' in a or 'vistoria' in a or 'perícia' in a or 'avaliação' in a or 'ensaio' in a: return 'Laudo / Perícia / Vistoria'
    if 'consultoria' in a or 'assessoria' in a or 'estudo' in a: return 'Consultoria / Estudo'
    if 'execu' in a or 'reforma' in a or 'manuten' in a or 'montagem' in a: return 'Execução / Obra'
    if 'fiscal' in a or 'supervis' in a: return 'Fiscalização / Supervisão'
    return 'Outros serviços técnicos'

rows=list(csv.DictReader(open(CAL,encoding='utf-8-sig')))
# nomes de coluna de faixa (piso_P20_R$, referencia_mediana_R$, teto_P80_R$)
hdr=rows[0].keys() if rows else []
kpiso=[h for h in hdr if h.startswith('piso')][0]
kref=[h for h in hdr if h.startswith('referencia')][0]
kteto=[h for h in hdr if h.startswith('teto')][0]
rows.sort(key=lambda r:(grupo(r['atividade'],r['unidade']), -int(r['n_total'])))

wb=Workbook()
HEAD=PatternFill('solid',fgColor='1F4E79'); HF=Font(color='FFFFFF',bold=True,size=10)
thin=Side(style='thin',color='BFBFBF'); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(wrap_text=True,vertical='top')

# ---- Leia-me ----
ws=wb.active; ws.title='Leia-me'
notas=[
 ['PLANILHA-MODELO — TABELA DE HONORÁRIOS DO SENGE/BA (estrutura proposta)'],[''],
 ['Fonte das faixas: ARTs CREA-BA 2022 (agregado, anonimizado).'],
 ['IMPORTANTE — a ART NÃO é honorário: valor_contrato pode ser valor de obra/contrato/declarado.'],
 ['As faixas (piso/referência/teto) são EVIDÊNCIA INDIRETA, para calibrar — não preço oficial.'],
 ['PODA DE OUTLIERS: faixas calculadas após remover 20% maiores e 20% menores valores de cada grupo.'],
 ['  piso = percentil 20 (aparado) · referência = mediana aparada · teto = percentil 80 (aparado).'],
 ['CONFIABILIDADE: "Baixa (provável mistura valor unitário/total)" sinaliza combos com teto muito acima'],
 ['  da mediana — provavelmente misturam valor unitário e valor total da obra; usar com cautela.'],
 ['NÃO PRECIFICÁVEL: registros sem valor/unidade (ex.: cargo-função) — ver aba "Precificabilidade".'],
 ['"honorario_referencia_R$" fica em branco ("Informação insuficiente para verificar") até a'],
 ['  pesquisa de preços (>=5 por item) e a definição de valor-hora (ver docs 04, 05, 10). NÃO inventar.'],
 ['Itens com n<5 foram suprimidos (LGPD). Aba "Tabela (modelo)": 1 linha por Atividade x Unidade.'],
]
for i,nt in enumerate(notas,1):
    c=ws.cell(i,1,nt[0])
    if i==1: c.font=Font(bold=True,size=13,color='1F4E79')
ws.column_dimensions['A'].width=112

# ---- Tabela (modelo) ----
ws=wb.create_sheet('Tabela (modelo)')
cols=['grupo_atividade','atividade','nivel_predominante','unidade_referencia','n_ART_2022',
      'piso_aparado_R$','referencia_mediana_R$','teto_aparado_R$','confiabilidade',
      'criterio_precificacao','honorario_referencia_R$','observacao']
for j,h in enumerate(cols,1):
    c=ws.cell(1,j,h); c.fill=HEAD; c.font=HF; c.alignment=WRAP; c.border=BORD
r=2
for d in rows:
    g=grupo(d['atividade'],d['unidade'])
    vals=[g,d['atividade'],d['nivel_predominante'],d['unidade'],int(d['n_total']),
          float(d[kpiso]),float(d[kref]),float(d[kteto]),d.get('confiabilidade',''),
          'Esforço técnico × '+d['unidade']+' (calibrar com referência aparada)',
          'Informação insuficiente para verificar',
          'Faixa observada em ART, com poda de 20% (evidência indireta; não é honorário)']
    for j,v in enumerate(vals,1):
        c=ws.cell(r,j,v); c.border=BORD; c.alignment=WRAP
        if j in (6,7,8): c.number_format='#,##0.00'
    r+=1
ws.freeze_panes='A2'
for j,w in enumerate([24,30,16,18,11,15,17,15,30,30,26,34],1): ws.column_dimensions[get_column_letter(j)].width=w
ws.auto_filter.ref=f'A1:{get_column_letter(len(cols))}{r-1}'

# ---- Precificabilidade ----
ws=wb.create_sheet('Precificabilidade')
pr=list(csv.DictReader(open(PREC,encoding='utf-8-sig')))
ws['A1']='Precificabilidade por tipo de ART'; ws['A1'].font=Font(bold=True,size=13,color='1F4E79')
ph=['tipo_art','total','precificaveis','nao_precificaveis','pct_precificavel']
for j,h in enumerate(ph,1):
    c=ws.cell(3,j,h); c.fill=HEAD; c.font=HF; c.border=BORD
for i,d in enumerate(pr,4):
    for j,k in enumerate(ph,1):
        c=ws.cell(i,j,(int(float(d[k])) if k!='pct_precificavel' and k!='tipo_art' else (d[k] if k=='tipo_art' else float(d[k])))); c.border=BORD
ws.cell(i+2,1,'Não precificável = registro sem valor e/ou sem unidade de medida (ex.: cargo-função, registros sem medida).')
for j,w in enumerate([42,12,14,16,16],1): ws.column_dimensions[get_column_letter(j)].width=w

# ---- Resumo ----
ws=wb.create_sheet('Resumo'); ws['A1']='Resumo'; ws['A1'].font=Font(bold=True,size=13,color='1F4E79')
res=[['Combinações Atividade×Unidade (n>=5)',len(rows)],['Atividades distintas',len({d['atividade'] for d in rows})],
     ['Ano-base',2022],['UF','BA (agregado)'],['Poda de outliers','20% em cada cauda'],
     ['Faixa','piso(P20) / referência(mediana) / teto(P80), aparados'],
     ['Honorário-referência','A definir por pesquisa de preços + valor-hora (não inventado)']]
for i,(k,v) in enumerate(res,3):
    ws.cell(i,1,k).font=Font(bold=True); ws.cell(i,2,v)
ws.column_dimensions['A'].width=42; ws.column_dimensions['B'].width=48

outp=os.path.join(BASE,'docs','entregaveis','PLANILHA_MODELO_HONORARIOS.xlsx')
os.makedirs(os.path.dirname(outp), exist_ok=True)
wb.save(outp)
print('WROTE',outp,'| itens:',len(rows),'| tipos precificabilidade:',len(pr))

