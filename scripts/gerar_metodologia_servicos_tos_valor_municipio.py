# -*- coding: utf-8 -*-
"""
Rodada TOS + Natureza do Valor + Municipio (SENGE/BA).
Fonte TOS: TABELA TOS - 2.xlsx (sheets 'Table 1' = hierarquia TOS; 'ARTs CREA 2022 (TOS)' = ARTs ja com CODIGO TOS).
NAO sobrescreve nada: gera apenas arquivos *_tos_valor_municipio* e relatorios.
Regras: nao soma linhas da mesma ART; valor da ART != honorario liquido; nao inventa codigo/municipio/IBGE.
"""
from __future__ import annotations
import csv, json, re, unicodedata, datetime
from collections import defaultdict, Counter
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parents[1]
REF = REPO / "assets" / "referencia"
PROCESSADO = REPO / "data" / "local" / "processado"
CAMADA_TOS_DIR = PROCESSADO / "camada_tos_2022"
TOS_XLSX = Path(r"C:\Users\adina\Meu Drive\SENGE\NOVO ARQUIVO\TABELA TOS - 2.xlsx")
CIDADES_XLSX = Path(r"C:\Users\adina\Meu Drive\SENGE\CidadessCalculo atualizado 07.11.2024.xlsx")
DIM_MUN = REF / "dim_municipios_bahia.csv"
INSUF = "Informação insuficiente para verificar"
INSUF_UNIDADE = "Informação insuficiente para verificar."
ABSURDO = 1_000_000_000.0  # > R$ 1 bilhao = implausivel

def strip_accents(s):
    s = unicodedata.normalize("NFD", "" if s is None else str(s))
    return "".join(c for c in s if unicodedata.category(c) != "Mn")

def norm_key(s):
    s = strip_accents(s).upper()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\s*[-/]\s*[A-Z]{2}$", "", s)  # remove sufixo UF
    return s.strip()

def unidade_segura(s):
    txt = str(s or "").strip()
    if not txt:
        return INSUF_UNIDADE
    key = norm_key(txt)
    if key in {"NAO INFORMADA", "NAO SELECIONADO", "SEM UNIDADE", "(NAO INFORMADA)", "(SEM UNIDADE)"}:
        return INSUF_UNIDADE
    return txt

# ---------- dicionario SENGE (mesmo do pipeline original, para comparabilidade) ----------
SVC_RULES = [
    (("RECEITUARIO",), "Receituario Agronomico", "Servico novo - Agronomia"),
    (("FOTOVOLTAIC", "MICROGERA", "GERACAO DE ENERGIA", "GERACAO DISTRIBUIDA", " SOLAR"),
        "Microgeracao / Fotovoltaica", "Servico novo - Energia"),
    (("ALVENARIA",), "Edificacao em Alvenaria", "Projetos Civis"),
    (("CONCRETO", "ESTRUTURA"), "Estrutura / Calculo Estrutural", "Calculo Estrutural"),
    (("BAIXA TENSAO", "INST.ELETR", "INSTALACAO ELETR", "INSTALACOES ELETR", "ELETR"),
        "Instalacoes Eletricas", "Instalacoes Eletricas e de Comunicacao"),
    (("HIDRO-SAN", "HIDROSSAN", "HIDRO SAN", "REDE HIDRO", "SANITARI", "ESGOTO", "DRENAGEM", "AGUA"),
        "Rede Hidrossanitaria / Saneamento", "Saneamento e Instalacoes Hidraulicas"),
    (("PAVIMENTA", "ESTRADA", "RODOVI"), "Pavimentacao / Estradas", "Estradas e Ruas"),
    (("TOPOGRAF", "GEORREF", "AGRIMENS", "GEODES", "CARTOGRAF"), "Topografia / Agrimensura", "Servico novo - Agrimensura"),
    (("AR CONDICIONADO", "CLIMATIZ", "VENTILA", "REFRIGERA"), "Climatizacao / Ventilacao", "Instalacoes Mecanicas"),
    (("INCENDIO",), "Prevencao de Incendio", "Instalacoes"),
    (("AMBIENT", "LICENCIAMENTO"), "Meio Ambiente / Licenciamento", "Servico novo - Ambiental"),
    (("SEGURANCA",), "Seguranca do Trabalho", "Servico novo - Seguranca"),
    (("LAUDO",), "Laudo Tecnico", "Pericias e Laudos"),
    (("VISTORIA", "PERICIA", "AVALIACAO"), "Vistoria / Pericia", "Pericias e Laudos"),
    (("DESEMPENHO DE CARGO", "CARGO TECNICO", "CARGO/FUNCAO", "CARGO-FUNCAO"),
        "Desempenho de Cargo / Funcao", "Cargo-Funcao"),
    (("CONSULTORIA", "ASSISTENCIA TECNICA"), "Consultoria Tecnica", "Consultoria"),
]
NAO_MAP = ("Nao mapeado", "Informacao insuficiente para verificar")

def map_servico(texto_norm, tipo_norm=""):
    if "RECEITUARIO" in tipo_norm or "RECEITUARIO" in texto_norm:
        return "Receituario Agronomico", "Servico novo - Agronomia"
    if "CARGO" in tipo_norm:
        return "Desempenho de Cargo / Funcao", "Cargo-Funcao"
    for kws, svc, grp in SVC_RULES:
        for kw in kws:
            if kw in texto_norm:
                return svc, grp
    return NAO_MAP

# ---------- 1) hierarquia TOS (Table 1) ----------
def load_tos_hierarchy():
    wb = openpyxl.load_workbook(TOS_XLSX, read_only=True, data_only=True)
    ws = wb["Table 1"]
    codemap = {}
    for r in ws.iter_rows(values_only=True):
        if not r or r[1] in (None, "Código", "Codigo"):
            continue
        cod = str(r[1]).strip()
        desc = ""
        for c in (r[2], r[3], r[4]):
            if c and str(c).strip():
                desc = str(c).strip(); break
        if cod and cod != "None":
            codemap[cod] = desc
    wb.close()
    return codemap

def resolve_tos(tos, codemap):
    tos = str(tos).strip()
    if not tos or tos == "None":
        return INSUF, INSUF, INSUF, INSUF
    segs = tos.split(".")
    grupo = codemap.get(segs[0], INSUF)
    sub = codemap.get(".".join(segs[:2]), INSUF) if len(segs) >= 2 else INSUF
    serv = codemap.get(tos, INSUF)
    # complemento = ultimo nivel descritivo (folha) quando difere do servico
    comp = serv
    return grupo, sub, serv, comp

# ---------- municipios conhecidos (validacao, sem IBGE oficial) ----------
def load_known_municipios():
    known = set()
    if DIM_MUN.exists():
        with open(DIM_MUN, encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                k = row.get("municipio_key") or norm_key(row.get("municipio_label", ""))
                if k: known.add(norm_key(k))
    try:
        wb = openpyxl.load_workbook(CIDADES_XLSX, read_only=True, data_only=True)
        ws = wb["Planilha1"]
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i < 3 or not r:
                continue
            c = r[0]
            if c and str(c).strip():
                known.add(norm_key(c))
        wb.close()
    except Exception:
        pass
    known.discard("")
    return known

# ---------- classificacao de confiabilidade (nivel ART) ----------
def classify(n, val, vary, ncods):
    has_val = val is not None and val > 0
    plausible = has_val and val < ABSURDO
    if not plausible:
        return "D", ("valor ausente/zerado" if not has_val else "valor implausivel (> R$ 1 bilhao)")
    if n == 1:
        return "A", ""
    if ncods <= 1 and vary == 0:
        return "B", "composto homogeneo (mesmo codigo, valor unico)"
    return "C", "composto ambiguo (multiplas atividades/valor unico replicado)"

# ---------- natureza do valor ----------
EXEC_NIV = {"EXECUCAO", "FISCALIZACAO", "DIRECAO DE OBRA", "SUPERVISAO", "COORDENACAO",
            "CONDUCAO DE EQUIPE", "EXECUCAO EM BIM", "DIRECAO DE OBRA EM BIM",
            "FISCALIZACAO EM BIM", "GESTAO", "GESTAO EM BIM"}
TEC_NIV = {"ELABORACAO", "ELABORACAO EM BIM", "CONSULTORIA", "CONSULTORIA EM BIM",
           "ASSESSORIA", "ASSESSORIA EM BIM", "ASSISTENCIA", "ORIENTACAO", "CONCEPCAO",
           "CONCEPCAO EM BIM", "CONDUCAO DE SERVICO TECNICO", "DIRECAO DE SERVICO TECNICO"}
OBRA_UNI = {"METRO QUADRADO", "METRO CUBICO", "METRO", "QUILOMETRO", "HECTARE"}
TEC_UNI = {"UNIDADE", "MES", "ANO", "DIA", "HORA", "HOMEM HORA"}

def natureza_base(val, niv_modal, uni_modal, niveis, unidades):
    if val is None or val <= 0:
        return "valor_inconsistente_ou_extremo", "alta", "valor ausente, zerado ou negativo"
    if val >= ABSURDO:
        return "valor_inconsistente_ou_extremo", "alta", "valor implausivel (>= R$ 1 bilhao)"
    if val <= 10:
        return "valor_simbolico_ou_taxa", "media", "valor <= R$ 10 (provavel taxa/valor simbolico)"
    nm = norm_key(niv_modal); um = norm_key(uni_modal)
    if nm in EXEC_NIV or um in OBRA_UNI:
        return "provavel_valor_obra_contrato", "media", f"nivel '{niv_modal}'/unidade '{uni_modal}' indicam execucao/obra"
    if nm in TEC_NIV and um in TEC_UNI:
        return "provavel_honorario_tecnico", "media", f"nivel '{niv_modal}' + unidade '{uni_modal}' indicam elaboracao tecnica"
    return "informacao_insuficiente", "baixa", f"combinacao nivel '{niv_modal}'/unidade '{uni_modal}' nao conclusiva"

def quantiles(vals):
    vals = sorted(vals); n = len(vals)
    def pct(p):
        if n == 0: return None
        k = (n - 1) * p; f = int(k); c = min(f + 1, n - 1)
        return vals[f] + (vals[c] - vals[f]) * (k - f)
    return pct(0.5), pct(0.25), pct(0.75), vals[0], vals[-1]

def main():
    print("Carregando hierarquia TOS...")
    codemap = load_tos_hierarchy()
    print("  codigos TOS na hierarquia:", len(codemap))
    known_mun = load_known_municipios()
    print("  municipios BA conhecidos (validacao):", len(known_mun))

    print("Lendo ARTs com TOS (pode levar ~40s)...")
    wb = openpyxl.load_workbook(TOS_XLSX, read_only=True, data_only=True)
    ws = wb["ARTs CREA 2022 (TOS)"]
    it = ws.iter_rows(values_only=True); next(it)
    arts = {}
    for r in it:
        if not r or r[0] is None:
            continue
        aid = str(r[0]).strip()
        if aid.endswith(".0"): aid = aid[:-2]
        tipo = str(r[1] or "")
        emissao = str(r[3] or "")
        cidade = str(r[4] or "")
        uf = str(r[5] or "").strip()
        titulo = str(r[6] or "")
        cod_niv = str(r[7] or "")
        nivel = str(r[8] or "")
        cod_ativ = str(r[9] or "").strip()
        ativ = str(r[10] or "")
        tos = str(r[11] or "").strip()
        desc_tos = str(r[12] or "")
        unidade = str(r[13] or "")
        try: vcontr = round(float(r[15]), 2)
        except Exception: vcontr = None
        m = re.search(r"/(\d{4})", emissao)
        ano = m.group(1) if m else ""
        a = arts.get(aid)
        if a is None:
            arts[aid] = {"n": 1, "val": vcontr, "vary": 0,
                         "cods": ({cod_ativ} if cod_ativ else set()),
                         "tos": [tos] if tos else [], "ativ": [ativ], "desc": [desc_tos],
                         "niv": [nivel], "uni": [unidade], "tipo": tipo, "titulo": titulo,
                         "cidade": cidade, "uf": uf, "ano": ano}
        else:
            a["n"] += 1
            if vcontr is not None:
                if a["val"] is None: a["val"] = vcontr
                elif abs((a["val"] or 0) - vcontr) > 0.005: a["vary"] = 1
            if cod_ativ: a["cods"].add(cod_ativ)
            if tos: a["tos"].append(tos)
            a["ativ"].append(ativ); a["desc"].append(desc_tos)
            a["niv"].append(nivel); a["uni"].append(unidade)
            if not a["ano"] and ano: a["ano"] = ano
    wb.close()
    print("  ARTs distintas (camada TOS):", len(arts))

    # ---------- monta linhas por ART ----------
    rows = []
    classe_count = Counter()
    nao_map_old = nao_map_new_tos = nao_map_new_senge = 0
    for aid, a in arts.items():
        n = a["n"]; val = a["val"]; vary = a["vary"]; ncods = max(1, len(a["cods"]))
        classe, motivo = classify(n, val, vary, ncods)
        classe_count[classe] += 1

        # TOS representativo (TOS mais frequente da ART)
        tos_rep = Counter(a["tos"]).most_common(1)[0][0] if a["tos"] else ""
        grupo_tos, sub_tos, serv_tos, comp_tos = resolve_tos(tos_rep, codemap)
        atividade_tos = Counter([d for d in a["desc"] if d]).most_common(1)[0][0] if any(a["desc"]) else INSUF

        # mapeamento SENGE NOVO (texto rico: grupo+sub+serv+atividade profissional)
        texto_rico = norm_key(" ".join([grupo_tos, sub_tos, serv_tos] + a["ativ"]))
        svc_new, grp_new = map_servico(texto_rico, norm_key(a["tipo"]))
        # mapeamento SENGE ANTIGO (so atividade profissional, p/ before/after no MESMO subconjunto)
        texto_old = norm_key(" ".join(a["ativ"]))
        svc_old, _ = map_servico(texto_old, norm_key(a["tipo"]))
        if svc_old == "Nao mapeado": nao_map_old += 1
        if not tos_rep: nao_map_new_tos += 1
        if svc_new == "Nao mapeado":
            nao_map_new_senge += 1
            servico_hon = "Nao mapeado (candidato a novo servico)"
            grupo_hon = grupo_tos if grupo_tos != INSUF else "Informacao insuficiente para verificar"
        else:
            servico_hon = svc_new; grupo_hon = grp_new

        nivel_conf_map = "exato_tos" if tos_rep else "nao_mapeado"
        fonte_map = "CODIGO TOS + Table 1 (TABELA TOS - 2.xlsx)" if tos_rep else INSUF

        # natureza do valor
        niv_modal = Counter([x for x in a["niv"] if x]).most_common(1)[0][0] if any(a["niv"]) else ""
        uni_modal = Counter([x for x in a["uni"] if x]).most_common(1)[0][0] if any(a["uni"]) else ""
        unidade_modal = unidade_segura(uni_modal)
        nat, nat_conf, nat_mot = natureza_base(val, niv_modal, uni_modal, a["niv"], a["uni"])

        # municipio
        mun_key = norm_key(a["cidade"])
        if not mun_key:
            mun_label, conf_mun, mot_mun = INSUF, "baixa", "cidade ausente"
        elif a["uf"] and a["uf"] != "BA":
            mun_label, conf_mun, mot_mun = mun_key.title(), "baixa", f"obra fora da BA (UF={a['uf']})"
        elif mun_key in known_mun:
            mun_label, conf_mun, mot_mun = mun_key.title(), "media", "nome casa lista BA conhecida (sem IBGE oficial)"
        else:
            mun_label, conf_mun, mot_mun = mun_key.title(), "baixa", "nome nao confirmado em lista BA local"

        usar = "sim" if classe == "A" else ("secundario" if classe == "B" else "nao")
        mot_excl = motivo if classe in ("C", "D") else ("uso secundario (composto homogeneo)" if classe == "B" else "")
        valor_repl = 1 if (n > 1 and vary == 0 and val is not None) else 0

        rows.append({
            "id_art": aid, "ano": a["ano"] or INSUF,
            "municipio": a["cidade"], "municipio_key": mun_key,
            "municipio_original": a["cidade"], "municipio_label": mun_label,
            "codigo_ibge": INSUF, "confiabilidade_municipio": conf_mun, "motivo_municipio": mot_mun, "uf": a["uf"],
            "atividade_original": INSUF, "atividade_key": servico_hon,
            "codigo_atividade": ";".join(sorted(a["cods"])) if a["cods"] else "",
            "codigo_tos": tos_rep or INSUF, "grupo_tos": grupo_tos, "subgrupo_tos": sub_tos,
            "servico_tos": serv_tos, "complemento_tos": comp_tos, "atividade_tos": atividade_tos,
            "servico_honorarios_padronizado": servico_hon, "grupo_servico_honorarios": grupo_hon,
            "servico_padronizado": servico_hon, "grupo_servico": grupo_hon,
            "nivel_confianca_mapeamento": nivel_conf_map, "fonte_mapeamento": fonte_map,
            "nivel_atividade": niv_modal, "unidade_medida": unidade_modal,
            "modalidade": a["titulo"], "formacao": a["titulo"],
            "valor_art": val if val is not None else "",
            "natureza_valor": nat, "confiabilidade_natureza_valor": nat_conf, "motivo_natureza_valor": nat_mot,
            "qtd_linhas_art": n, "qtd_atividades_art": len(set(a["tos"])) if a["tos"] else len(a["niv"]),
            "qtd_codigos_atividade_art": len(a["cods"]),
            "valor_replicado_linhas": valor_repl,
            "classe_confiabilidade": classe, "usar_calculo_monetario": usar,
            "motivo_exclusao_calculo": mot_excl,
        })

    total = len(rows)
    print("Classes:", dict(classe_count))

    # ---------- 2-pass: outlier IQR por servico (so candidatos a honorario) ----------
    vals_by_svc_unidade = defaultdict(list)
    for r in rows:
        if (r["classe_confiabilidade"] == "A" and r["natureza_valor"] == "provavel_honorario_tecnico"
                and isinstance(r["valor_art"], (int, float))):
            vals_by_svc_unidade[(r["servico_honorarios_padronizado"], r["unidade_medida"])].append(r["valor_art"])
    iqr_lim = {}
    for svc_unidade, vs in vals_by_svc_unidade.items():
        if len(vs) >= 30:
            med, q1, q3, mn, mx = quantiles(vs)
            iqr_lim[svc_unidade] = q3 + 3 * (q3 - q1)
    n_out = 0
    for r in rows:
        if (r["natureza_valor"] == "provavel_honorario_tecnico"
                and isinstance(r["valor_art"], (int, float))):
            lim = iqr_lim.get((r["servico_honorarios_padronizado"], r["unidade_medida"]))
            if lim is not None and r["valor_art"] > lim:
                r["natureza_valor"] = "valor_inconsistente_ou_extremo"
                r["confiabilidade_natureza_valor"] = "media"
                r["motivo_natureza_valor"] = "outlier IQR (> Q3 + 3*IQR) no servico e unidade"
                n_out += 1
    print("  reclass. outlier IQR:", n_out)

    # ---------- escrita base ----------
    base_fields = ["id_art","ano","municipio","municipio_key","municipio_original","municipio_label",
        "codigo_ibge","confiabilidade_municipio","motivo_municipio","uf","atividade_original","atividade_key",
        "codigo_atividade","codigo_tos","grupo_tos","subgrupo_tos","servico_tos","complemento_tos","atividade_tos",
        "servico_honorarios_padronizado","grupo_servico_honorarios","servico_padronizado","grupo_servico",
        "nivel_confianca_mapeamento","fonte_mapeamento","nivel_atividade","unidade_medida","modalidade","formacao",
        "valor_art","natureza_valor","confiabilidade_natureza_valor","motivo_natureza_valor",
        "qtd_linhas_art","qtd_atividades_art","qtd_codigos_atividade_art","valor_replicado_linhas",
        "classe_confiabilidade","usar_calculo_monetario","motivo_exclusao_calculo"]
    with open(REPO / "base_servicos_tos_valor_municipio.csv", "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=base_fields, extrasaction="ignore")
        w.writeheader()
        for r in rows: w.writerow(r)

    # ---------- agregado confiavel: A + mapeado + honorario + n>=5 ----------
    NAT_SEGURA = {"provavel_honorario_tecnico"}
    agg = defaultdict(lambda: {"vals": [], "anos": set(), "muns": set(), "mods": set(),
                               "grupo_tos": set(), "sub_tos": set(), "serv_tos": set(), "grupo_hon": ""})
    for r in rows:
        if (r["classe_confiabilidade"] == "A"
                and r["servico_honorarios_padronizado"] not in ("Nao mapeado", "Nao mapeado (candidato a novo servico)")
                and r["natureza_valor"] in NAT_SEGURA
                and isinstance(r["valor_art"], (int, float))):
            k = (r["servico_honorarios_padronizado"], r["unidade_medida"])
            g = agg[k]
            g["vals"].append(r["valor_art"]); g["anos"].add(r["ano"]); g["muns"].add(r["municipio_key"])
            g["mods"].add(r["modalidade"]); g["grupo_tos"].add(r["grupo_tos"]); g["sub_tos"].add(r["subgrupo_tos"])
            g["serv_tos"].add(r["servico_tos"]); g["grupo_hon"] = r["grupo_servico_honorarios"]
    serv_ok = serv_insuf = 0
    REF.mkdir(parents=True, exist_ok=True)
    with open(REF / "agregado_servicos_tos_classe_a_valor_confiavel.csv", "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["servico_honorarios_padronizado","unidade_medida","grupo_servico_honorarios","grupo_tos","subgrupo_tos",
                    "servico_tos","n_arts","mediana_valor","q1","q3","iqr","min","max","ano_min","ano_max",
                    "municipios_distintos","modalidades_distintas","natureza_valor","confiabilidade","observacao"])
        def topset(s, k=3):
            vals = sorted(x for x in s if x and x != INSUF)
            return "; ".join(vals[:k]) + (" …" if len(vals) > k else "") if vals else INSUF
        for (svc, unidade), g in sorted(agg.items(), key=lambda kv: -len(kv[1]["vals"])):
            n = len(g["vals"]); anos = sorted(x for x in g["anos"] if x and x[0].isdigit())
            if n < 5:
                serv_insuf += 1
                w.writerow([svc, unidade, g["grupo_hon"], topset(g["grupo_tos"]), topset(g["sub_tos"]), topset(g["serv_tos"]),
                            n, INSUF, INSUF, INSUF, INSUF, INSUF, INSUF,
                            (anos[0] if anos else ""), (anos[-1] if anos else ""),
                            len(g["muns"]), len(g["mods"]), "provavel_honorario_tecnico", "insuficiente",
                            "n < 5: sem base estatistica para referencia monetaria"])
            else:
                serv_ok += 1
                med, q1, q3, mn, mx = quantiles(g["vals"])
                conf = "alta" if n >= 30 else "media"
                w.writerow([svc, unidade, g["grupo_hon"], topset(g["grupo_tos"]), topset(g["sub_tos"]), topset(g["serv_tos"]),
                            n, round(med,2), round(q1,2), round(q3,2), round(q3-q1,2), round(mn,2), round(mx,2),
                            (anos[0] if anos else ""), (anos[-1] if anos else ""),
                            len(g["muns"]), len(g["mods"]), "provavel_honorario_tecnico", conf,
                            "valor declarado em ART (filtrado p/ provavel honorario); nao e honorario liquido"])

    # ---------- resumo natureza do valor ----------
    nat_agg = defaultdict(list)
    for r in rows:
        if isinstance(r["valor_art"], (int, float)):
            nat_agg[r["natureza_valor"]].append(r["valor_art"])
        else:
            nat_agg[r["natureza_valor"]]  # garante chave
    nat_count = Counter(r["natureza_valor"] for r in rows)
    with open(REF / "resumo_natureza_valor.csv", "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["natureza_valor","quantidade_ARTs","percentual","mediana","q1","q3","observacao"])
        OBS = {
            "provavel_honorario_tecnico": "elaboracao/projeto/consultoria com unidade tecnica",
            "provavel_valor_obra_contrato": "execucao/fiscalizacao ou unidade de obra; valor de contrato",
            "valor_simbolico_ou_taxa": "valor <= R$ 10",
            "valor_inconsistente_ou_extremo": "ausente/zerado, >= R$ 1 bilhao ou outlier IQR",
            "informacao_insuficiente": "combinacao nivel/unidade nao conclusiva",
        }
        for nat in ["provavel_honorario_tecnico","provavel_valor_obra_contrato","valor_simbolico_ou_taxa",
                    "valor_inconsistente_ou_extremo","informacao_insuficiente"]:
            q = nat_count.get(nat, 0); vs = [v for v in nat_agg.get(nat, []) if isinstance(v,(int,float))]
            if vs:
                med, q1, q3, _, _ = quantiles(vs)
                w.writerow([nat, q, f"{round(100*q/total,1)}%", round(med,2), round(q1,2), round(q3,2), OBS[nat]])
            else:
                w.writerow([nat, q, f"{round(100*q/total,1)}%", INSUF, INSUF, INSUF, OBS[nat]])

    # ---------- diagnostico padronizacao municipios ----------
    mun_diag = defaultdict(lambda: {"orig": set(), "n": 0, "uf": set(), "conf": ""})
    for r in rows:
        k = r["municipio_key"] or "(vazio)"
        d = mun_diag[k]; d["orig"].add(r["municipio_original"]); d["n"] += 1
        d["uf"].add(r["uf"]); d["conf"] = r["confiabilidade_municipio"]
    with open(REF / "diagnostico_padronizacao_municipios.csv", "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["municipio_key","municipio_label","codigo_ibge","grafias_distintas","n_arts","uf",
                    "confiabilidade_municipio","exemplos_grafias"])
        for k, d in sorted(mun_diag.items(), key=lambda kv: -kv[1]["n"]):
            ex = "; ".join(sorted(x for x in d["orig"] if x)[:4])
            w.writerow([k, k.title() if k != "(vazio)" else INSUF, INSUF, len(d["orig"]), d["n"],
                        "; ".join(sorted(x for x in d["uf"] if x)), d["conf"], ex])
    n_mun_total = len([k for k in mun_diag if k != "(vazio)"])
    n_mun_media = len([1 for k, d in mun_diag.items() if d["conf"] == "media"])
    n_mun_baixa = len([1 for k, d in mun_diag.items() if d["conf"] == "baixa"])

    # ---------- JSON para dashboard ----------
    svc_list, svc_idx = [], {}; unidade_list, unidade_idx = [], {}; mun_list, mun_idx = [], {}; ano_list, ano_idx = [], {}
    grp_of, nat_list, nat_idx, grptos_list, grptos_idx = {}, [], {}, [], {}
    def gi(lst, idx, key):
        if key not in idx: idx[key] = len(lst); lst.append(key)
        return idx[key]
    CL = {"A":0,"B":1,"C":2,"D":3}
    cA = {"s": [], "u": [], "m": [], "a": [], "v": [], "nat": [], "gt": []}
    agg2 = defaultdict(lambda: [0,0])  # (cl,svc,unidade,ano,mun,nat,gt)->[n,ativ]
    for r in rows:
        s = gi(svc_list, svc_idx, r["servico_honorarios_padronizado"]); grp_of[s] = r["grupo_servico_honorarios"]
        u = gi(unidade_list, unidade_idx, unidade_segura(r["unidade_medida"]))
        m = gi(mun_list, mun_idx, r["municipio_key"] or "(nao informado)")
        an = gi(ano_list, ano_idx, r["ano"])
        nt = gi(nat_list, nat_idx, r["natureza_valor"])
        gt = gi(grptos_list, grptos_idx, r["grupo_tos"])
        ci = CL[r["classe_confiabilidade"]]
        cell = agg2[(ci,s,u,an,m,nt,gt)]; cell[0]+=1; cell[1]+=r["qtd_atividades_art"]
        if r["classe_confiabilidade"]=="A" and isinstance(r["valor_art"],(int,float)):
            cA["s"].append(s); cA["u"].append(u); cA["m"].append(m); cA["a"].append(an); cA["v"].append(round(r["valor_art"],2)); cA["nat"].append(nt); cA["gt"].append(gt)
    data = {
        "gerado_em": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "fonte": "TABELA TOS - 2.xlsx, aba 'ARTs CREA 2022 (TOS)' (camada TOS = subconjunto da base 2022)",
        "total_arts": total,
        "universo_total_2022": 230928,
        "classe_count": dict(classe_count),
        "nao_mapeado": {"old_keyword": nao_map_old, "new_tos": nao_map_new_tos, "new_senge": nao_map_new_senge,
                        "total_subconjunto": total},
        "natureza_count": {k: nat_count.get(k,0) for k in
            ["provavel_honorario_tecnico","provavel_valor_obra_contrato","valor_simbolico_ou_taxa",
             "valor_inconsistente_ou_extremo","informacao_insuficiente"]},
        "servicos": svc_list, "grupo_de_servico": [grp_of.get(i,"") for i in range(len(svc_list))],
        "unidades": unidade_list,
        "municipios": mun_list, "anos": ano_list, "naturezas": nat_list, "grupos_tos": grptos_list,
        "classeA": cA, "agg": [[k[0],k[1],k[2],k[3],k[4],k[5],k[6],v[0],v[1]] for k,v in agg2.items()],
    }
    CAMADA_TOS_DIR.mkdir(parents=True, exist_ok=True)
    (CAMADA_TOS_DIR / "dados_tos_valor_municipio_2022_tos_local.json").write_text(
        json.dumps(data, ensure_ascii=False),
        encoding="utf-8",
    )

    # ---------- STATS para relatorios ----------
    stats = {
        "total_subconjunto_tos": total, "classe_count": dict(classe_count),
        "nao_map_old": nao_map_old, "nao_map_new_tos": nao_map_new_tos, "nao_map_new_senge": nao_map_new_senge,
        "natureza": {k: nat_count.get(k,0) for k in
            ["provavel_honorario_tecnico","provavel_valor_obra_contrato","valor_simbolico_ou_taxa",
             "valor_inconsistente_ou_extremo","informacao_insuficiente"]},
        "serv_confiavel_n>=5": serv_ok, "serv_confiavel_n<5": serv_insuf,
        "outliers_iqr": n_out, "n_servicos": len(svc_list),
        "mun_total": n_mun_total, "mun_media": n_mun_media, "mun_baixa": n_mun_baixa,
    }
    print("STATS_JSON_BEGIN")
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    print("STATS_JSON_END")

if __name__ == "__main__":
    main()

